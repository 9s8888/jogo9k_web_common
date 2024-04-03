# 运行方式 python [脚本] -out [文件输出目录] [xlsx文件路径]
import sys
import os
import math
import argparse
try:
    import jmutil.logger as logger
    LOG_RED = logger.CC.RED
    LOG_GREEN = logger.CC.GREEN
    LOG_END = logger.CC.END
    USE_LOGGER = True
except ImportError:
    LOG_RED = ''
    LOG_GREEN = ''
    LOG_END = ''
    USE_LOGGER = False

import warnings
warnings.simplefilter('ignore')

VERSION = '2.1'
ERRORS = []
KEYWORD_EXPORT = 'export'
TRUE_WORDS = ('true', '是', 'yes')
FALSE_WORDS = ('false', '否', 'no')

FILTER = None
FILTER_CLIENT = 'client'
FILTER_SERVER = 'server'


def error(msg, color=LOG_RED):
    desc = color+msg+LOG_END
    if USE_LOGGER:
        logger.log_error(msg, color=color, report=True)
    else:
        print(desc, file=sys.stderr)
    ERRORS.append(desc)


def write_json(name, dst, content):
    if not os.path.exists(dst):
        os.makedirs(dst)
    filename = os.path.join(dst, name+'.json')
    import json
    result = json.dumps(content, ensure_ascii=False, indent=4, sort_keys=True)
    f = open(filename, 'wt', encoding='utf-8')
    f.write(result)
    f.close()
    print('已导出：', filename)


def read_bool(val):
    if isinstance(val, str):
        if val.strip().lower() in TRUE_WORDS:
            return True
        else:
            return False
    elif val:
        return True
    else:
        return False


def read_int(val):
    if type(val) is str:
        return int(val, 0)
    else:
        return int(val)


def read_float(val):
    return float(val)


def read_str(val):
    if isinstance(val, str):
        return val
    else:
        return str(val)


def read_array(val):
    if isinstance(val, str):
        import json
        ret_list = json.loads(val)
    else:
        ret_list = [val]
    return ret_list


def read_auto(val):
    if isinstance(val, str):
        try:
            v = float(val)
            if math.floor(v) == v:
                return int(v)
            else:
                return v
        except ValueError:
            if val.startswith('[') and val.endswith(']'):
                return read_array(val)
            elif val.lower() in TRUE_WORDS:
                return True
            elif val.lower() in FALSE_WORDS:
                return False
            else:
                return val
    else:
        return val


type_reader = {
    'bool': read_bool,
    'int': read_int,
    'float': read_float,
    'str': read_str,
    'array': read_array,
    'json': read_array,
    'auto': read_auto,
}


def get_valid_type(type_str):
    if type_str.lower() in type_reader:
        return type_str.lower()
    else:
        return None


def read_value(typ, value):
    if value is None:
        return None
    if type(value) is str and value == '':
        return None

    if typ in type_reader:
        return type_reader[typ](value)
    else:
        return None


class Head(object):
    """
    表头列对象
    """
    def __init__(self):
        self.sheet = None
        self.desc = None
        self.name = None
        self.type = None
        self.default = None
        self.is_primary_key = False
        self.is_group_key = False
        self.is_not_null = False
        self.is_export = False
        self.offset_col = 0
        self.comment = None
        self.filter = None

    def setup(self, sheet, col):
        """
        读取表头列配置
        """
        self.sheet = sheet
        self.desc = sheet.cell(column=col, row=1).value
        self.comment = sheet.cell(column=col, row=1).comment
        if self.comment:
            self.comment = self.comment.text
        name = sheet.cell(column=col, row=2).value
        type = sheet.cell(column=col, row=3).value

        if name and isinstance(name, str) and name != '':
            # 处理服务器，客户端过滤
            if name.endswith('&'):
                name = name[:-1]
                self.filter = FILTER_CLIENT
            if name.endswith('@'):
                name = name[:-1]
                self.filter = FILTER_SERVER
            # 处理字段标记
            if name.startswith('*'):
                self.name = name[1:]
                self.is_primary_key = True
            elif name.startswith('#'):
                self.name = name[1:]
                self.is_group_key = True
            elif name.startswith('!'):
                self.name = name[1:]
                self.is_not_null = True
            else:
                self.name = name
                if self.name.lower() == KEYWORD_EXPORT:
                    self.is_export = True
        else:
            return False
        if type and isinstance(type, str) and type != '':
            assign = type.split('=')
            self.type = get_valid_type(assign[0])
            if self.type:
                if len(assign) > 1:
                    self.default = read_value(self.type, assign[1])
            else:
                return False
        self.offset_col = col
        return True

    def read(self, row):
        val = self.sheet.cell(column=self.offset_col, row=row).value
        ret = read_value(self.type, val)
        if ret is None and self.default:
            ret = self.default
        return ret

    def print(self):
        """
        debug usage
        """
        ret = '[%s|%s|%s' % (self.desc, self.name, self.type)
        if self.default is not None:
            ret += '=%s' % self.default
        if self.is_primary_key or self.is_group_key or self.is_not_null or self.is_export:
            ret += '('
            if self.is_primary_key:
                ret += '*'
            if self.is_group_key:
                ret += '#'
            if self.is_not_null:
                ret += '!'
            if self.is_export:
                ret += '@'
            ret += ')'
        ret += ']'
        print(ret)


def read_heads(sheet):
    max_col = sheet.max_column
    max_row = sheet.max_row
    if max_row >= 4 and max_col >= 1:
        ret_list = []
        for col in range(1, max_col+1):
            head = Head()
            if head.setup(sheet, col):
                ret_list.append(head)
        if len(ret_list) > 0:
            return ret_list
        else:
            return None
    else:
        return None


def read_content(out, heads, sheet_name):
    # prepare
    sheet = heads[0].sheet
    root = out
    primary_head = None
    group_head = None
    for h in heads:
        if h.is_primary_key:
            primary_head = h
        if h.is_group_key:
            group_head = h

    if primary_head is None:
        error('表[%s] 没有主键！' % heads[0].sheet.title)
        return False

    # read
    for row in range(4, sheet.max_row+1):
        data = {}
        is_export = True
        need_export = True
        skip_causes = []
        for h in heads:
            val = None

            # 读取数据
            try:
                val = h.read(row)
            except ValueError as e:
                error('数据错误: '+str(e))
                error('位置{表[%s] 字段[%s(%s):%s] 行[%d]}' % (sheet_name, h.desc, h.name, h.type, row), LOG_GREEN)

            # 过滤判断
            if h.filter is not None and FILTER is not None and h.filter != FILTER:
                continue

            # 检查
            if h.is_not_null and val is None:
                skip_causes.append('非空字段空缺: %s(%s)' % (h.name, h.desc))
                is_export = False
                break
            if h.is_primary_key and val is None:
                skip_causes.append('主键字段空缺: %s(%s)' % (h.name, h.desc))
                is_export = False
                break
            if h.is_group_key and val is None:
                skip_causes.append('分组字段空缺: %s(%s)' % (h.name, h.desc))
                is_export = False
                break

            if h.is_export:
                if not val:
                    is_export = False
                    need_export = False
                    break
            else:
                if val is not None:
                    data[h.name] = val

        if need_export:
            if is_export:
                key = data[primary_head.name]
                if group_head:
                    group = data[group_head.name]
                    if group not in root:
                        root[group] = {}
                    root[group][key] = data
                else:
                    if key in root:
                        error('主键重复: %s' % key)
                        error('位置{表[%s] 行[%d]}' % (sheet_name, row), LOG_GREEN)
                    root[key] = data
            else:
                for err in skip_causes:
                    error(err)
                    error('位置{表[%s] 行[%d]}' % (sheet_name, row), LOG_GREEN)

    return True


# 导出表格
def export_book(src, dst):
    print('开始导出工作簿:', src)
    if dst is None:
        dst = os.path.dirname(src)

    from openpyxl import load_workbook
    try:
        book = load_workbook(filename=src, data_only=True)
    except IOError:
        error('未能打开文件: %s' % src)
        return

    # 构建导出映射（用来支持表格合并）
    export_map = {}
    for sheet in book:
        if not sheet.title.startswith('#'):
            title = sheet.title.split('|')[0]
            if title.endswith('&'):
                title = title[:-1]
                if FILTER == FILTER_SERVER:
                    continue
            if title.endswith('@'):
                title = title[:-1]
                if FILTER == FILTER_CLIENT:
                    continue
            if title in export_map:
                export_map[title].append(sheet)
            else:
                export_map[title] = [sheet]

    # 导出表格
    for target, sheet_list in export_map.items():
        content = {}
        err_flag = False
        ext_flag = False
        for sheet in sheet_list:
            heads = read_heads(sheet)
            if heads:
                print('开始导出数据页:', sheet.title)
                ext_flag = True
                if not read_content(content, heads, sheet.title):
                    err_flag = True
                    break
        if not err_flag and ext_flag:
            write_json(target, dst, content)


def main(inputs, output=None, filter=None):
    """
    主要工作函数
    :param inputs: 输入
    :param output: 输出
    :param filter: ‘client’ | ‘server’ | None 表示只导出一部分数据
    """
    print('==========================================')
    global FILTER
    FILTER = filter
    if len(inputs) > 0:
        for item in inputs:
            if os.path.exists(item):
                if os.path.isdir(item):
                    for root, dirs, files in os.walk(item):
                        for fi in files:
                            if fi.endswith('.xlsx') and not fi.startswith('~$'):
                                export_book(os.path.join(root, fi), output)
                else:
                    export_book(item, output)
    print('==========================================')
    print('导表结果:')
    print('------------------------------------------')
    if len(ERRORS) > 0:
        for e in ERRORS:
            print(e, file=sys.stderr)
        print('')
        return 1
    else:
        print('所有表格已成功导出！')
        print('')
        return 0


# 作为独立工具使用
if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.description = 'JM导表工具 V%s' % VERSION
    parser.add_argument('inputs', nargs='*', help='源目录；Excel表所在的目录（支持多个）')
    parser.add_argument('-out', dest='output', help='输出目录（默认同源目录）')
    parser.add_argument('-c', action='store_const', dest='filter', const=FILTER_CLIENT, help='只输出客户端数据，对应&标记（默认全部输出）')
    parser.add_argument('-s', action='store_const', dest='filter', const=FILTER_SERVER, help='只输出服务器端数据，对应@标记（默认全部输出）')
    args = parser.parse_args()
    # 开始工作
    if len(args.inputs) > 0:
        code = main(args.inputs, args.output, filter=args.filter)
        sys.exit(code)
    else:
        parser.print_usage()
        sys.exit(1)
